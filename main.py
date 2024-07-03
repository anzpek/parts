import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from utils.file_selector import select_file, show_popup
from utils.popup_selector import select_brand, select_execution_mode
from manufacturers.bmw import scrape_bmw
from manufacturers.toyota import scrape_toyota
from manufacturers.hyundai_kia import scrape_hyundai_kia
from manufacturers.chevrolet import scrape_chevrolet
from manufacturers.hyundai_kia import select_hyundai_manufacturer, select_hyundai_search_mode
from manufacturers.chevrolet import select_chevrolet_model, select_chevrolet_search_type
from manufacturers.tesla import select_tesla_country, select_tesla_model, select_tesla_product, select_tesla_catalog, scrape_tesla
from manufacturers.benz import select_benz_model, scrape_benz
from manufacturers.audi import select_audi_year, select_audi_model, scrape_audi
from manufacturers.volkswagen import select_volkswagen_year, select_volkswagen_model, scrape_volkswagen
from manufacturers.jeep import select_jeep_model, scrape_jeep
from manufacturers.porsche import scrape_porsche
from manufacturers.lexus import scrape_lexus
from manufacturers.nissan import select_nissan_model, select_nissan_search_type, scrape_nissan
from manufacturers.kg import select_kg_model, scrape_kg
from manufacturers.honda import scrape_honda
from manufacturers.peugeot import scrape_peugeot
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import signal
import atexit

driver = None

def initialize_driver(execution_mode):
    global driver
    options = webdriver.ChromeOptions()
    if execution_mode == "headless":
        options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    return driver

def save_to_excel(results, file_path, sheet_name):
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        show_popup("엑셀 파일을 열 수 없습니다. 파일이 열려 있는지 확인하고 다시 실행하세요.")
        sys.exit(1)
    
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]

    # 기존 데이터 지우기
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=8, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # 새로운 데이터 추가
    for r_idx, result in enumerate(results, start=2):
        ws.cell(row=r_idx, column=2, value=result[0])
        ws.cell(row=r_idx, column=3, value=result[1])
        ws.cell(row=r_idx, column=4, value=result[2])
        ws.cell(row=r_idx, column=5, value=result[3])
        ws.cell(row=r_idx, column=6, value=result[4])
        ws.cell(row=r_idx, column=7, value=result[5])
        ws.cell(row=r_idx, column=8, value=result[6])

    try:
        wb.save(file_path)
        print(f"작업 완료. 결과가 '{sheet_name}' 시트에 저장되었습니다.")
    except Exception as e:
        show_popup("엑셀 파일을 저장할 수 없습니다. 파일이 열려 있는지 확인하고 다시 실행하세요.")
        sys.exit(1)

def handle_exit(sig, frame):
    global driver
    if driver:
        driver.quit()
    sys.exit(0)

def main():
    signal.signal(signal.SIGINT, handle_exit)
    signal.signal(signal.SIGTERM, handle_exit)
    atexit.register(handle_exit, None, None)  # 프로그램 종료 시 handle_exit 호출

    file_path = select_file()
    sheet_name = 'Sheet1'

    try:
        wb = load_workbook(file_path)
        wb.close()
    except Exception as e:
        show_popup(f"엑셀 파일을 열 수 없습니다. 파일이 열려 있는지 확인하고 다시 실행하세요.\n에러: {str(e)}")
        sys.exit(1)

    df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl', dtype=str)
    df = df.dropna(subset=[df.columns[0]])

    execution_mode = select_execution_mode()
    driver = None

    try:
        driver = initialize_driver(execution_mode)
        brand = select_brand()
        results = []

        try:
            if brand == "BMW":
                results = scrape_bmw(driver, df)
            elif brand == "Toyota":
                results = scrape_toyota(driver, df)
            elif brand == "Hyundai_Kia":
                search_mode = select_hyundai_search_mode()
                manufacturer = select_hyundai_manufacturer()
                results = scrape_hyundai_kia(driver, df, search_mode, manufacturer)
            elif brand == "Chevrolet":
                chevrolet_model = select_chevrolet_model()
                chevrolet_search_type = select_chevrolet_search_type()
                results = scrape_chevrolet(driver, df, chevrolet_model, chevrolet_search_type)
            elif brand == "Tesla":
                country = select_tesla_country()
                product = select_tesla_product()
                model = select_tesla_model()
                catalog = select_tesla_catalog(model)
                results = scrape_tesla(driver, df, country, product, model, catalog)
            elif brand == "Benz":
                model = select_benz_model()
                results = scrape_benz(driver, df, model)
            elif brand == "Audi":
                year = select_audi_year()
                model = select_audi_model(year)
                results = scrape_audi(driver, df, year, model)
            elif brand == "Volkswagen":
                year = select_volkswagen_year()
                model = select_volkswagen_model(year)
                results = scrape_volkswagen(driver, df, year, model)
            elif brand == "Jeep":
                model = select_jeep_model()
                results = scrape_jeep(driver, df, model)
            elif brand == "Porsche":
                results = scrape_porsche(driver, df)
            elif brand == "Lexus":
                results = scrape_lexus(driver, df)    
            elif brand == "Nissan":
                model = select_nissan_model()
                search_type = select_nissan_search_type()
                results = scrape_nissan(driver, df, model, search_type) 
            elif brand == "KG(쌍용)":
                model = select_kg_model()
                results = scrape_kg(driver, df, model) 
            elif brand == "Honda":
                results = scrape_honda(driver, df) 
            elif brand == "Peugeot":
                results = scrape_peugeot(driver, df) 
        except Exception as e:
            show_popup(f"스크래핑 도중 오류가 발생했습니다: {str(e)}")

        if results:
            save_to_excel(results, file_path, sheet_name)
            show_popup("부품 찾기 매크로가 성공적으로 실행을 완료했습니다.", file_path)
        else:
            show_popup("결과가 없습니다. 스크래핑 도중 오류가 발생했을 수 있습니다.")
    finally:
        if driver:
            driver.quit()

if __name__ == "__main__":
    main()
